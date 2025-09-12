from flask import render_template, request, redirect, url_for, flash, abort, Response, send_file
from flask_login import login_required, current_user
from sqlalchemy import or_
from ..utils.decorators import role_required
from ..extensions import db
from . import professors_bp
from .forms import GradeForm
from sqlalchemy.exc import IntegrityError
from collections import defaultdict
from ..models import Instrument  # garanta o import
from sqlalchemy import func
from io import BytesIO
import csv, io, datetime
from sqlalchemy import func, case, or_
from flask import send_file
from ..models import Group, GroupStudent, Student, Offering, Campus, User, GroupAssessment, Instrument
from ..models import Offering  # certifique-se de importar Offering
from sqlalchemy.sql import func
from sqlalchemy.orm import joinedload
from datetime import datetime          # <— importante: classe datetime
from io import StringIO, BytesIO       # <— para CSV/XLSX em memória
import csv

from datetime import datetime  # se ainda não tiver

from app.models import (
    db, Offering, Student, Group, User, GroupAssessment, Instrument, GroupStudent
)

from decimal import Decimal
from app.services.grades import upsert_assessment, get_assessment_score
# app/admin/routes.py
from app.models import Instrument

# Aliases compatíveis com os dois jeitos que você usou no Enum
INST_RI    = getattr(Instrument, "RI", None)  or getattr(Instrument, "relatorio_i")
INST_RII   = getattr(Instrument, "RII", None) or getattr(Instrument, "relatorio_ii")
INST_PAPER = getattr(Instrument, "PAPER", None) or getattr(Instrument, "paper")


def _stamp():
    return datetime.now().strftime("%Y%m%d-%H%M%S")

def _ensure_owns_group(group_id: int) -> Group:
    g = Group.query.get_or_404(group_id)
    if g.orientador_user_id != current_user.id:
        abort(403)
    return g

def _fmt(n):
    """Formata nota com 2 casas e vírgula; vazio se None."""
    if n is None:
        return ""
    try:
        return f"{float(n):.2f}".replace(".", ",")
    except Exception:
        return str(n)

def _assessments_by_instrument(group_id: int):
    rows = GroupAssessment.query.filter_by(group_id=group_id).all()
    # chave = Enum Instrument; valor = linha
    return {row.instrument: row for row in rows}

def _upsert_grade(group_id: int, instrument, score):
    """
    Upsert de uma nota (0–10) para um instrumento específico.
    Se 'score' for None -> remove a linha daquele instrumento.
    """
    row = GroupAssessment.query.filter_by(group_id=group_id, instrument=instrument).first()
    if score is None:
        if row:
            db.session.delete(row)
        return
    # score preenchido -> cria/atualiza
    if row:
        row.score = score
        row.entered_by_user_id = current_user.id
    else:
        db.session.add(GroupAssessment(
            group_id=group_id,
            instrument=instrument,
            score=score,
            entered_by_user_id=current_user.id
        ))

@professors_bp.route("/", methods=["GET"])
@login_required
@role_required("professor")
def dashboard():
    return render_template("professors/dashboard.html")

@professors_bp.route("/offerings", methods=["GET"])
@login_required
@role_required("professor")
def offerings_list():
    q = (request.args.get("q") or "").strip()
    base = Offering.query.filter(Offering.professor_id == current_user.id)
    if q:
        like = f"%{q}%"
        base = base.filter(or_(Offering.code.ilike(like),
                               Offering.description.ilike(like)))
    offerings = base.order_by(Offering.code.asc()).all()
    return render_template("professors/offerings_list.html", offerings=offerings, q=q)

@professors_bp.route("/groups")
@login_required
@role_required("professor")
def groups_list():    

    groups = (Group.query
              .filter_by(orientador_user_id=current_user.id)
              .order_by(Group.id.asc())
              .all())

    group_ids = [g.id for g in groups] or [0]
    rows = GroupAssessment.query.filter(
        GroupAssessment.group_id.in_(group_ids)
    ).all()

    scores = {gid: {"ri": None, "rii": None, "paper": None} for gid in group_ids}
    for r in rows:
        if r.instrument == Instrument.RELATORIO_I:
            scores[r.group_id]["ri"] = r.score
        elif r.instrument == Instrument.RELATORIO_II:
            scores[r.group_id]["rii"] = r.score
        elif r.instrument == Instrument.PAPER:
            scores[r.group_id]["paper"] = r.score

    #return render_template("professors/dashboard.html", groups=groups, scores=scores)
    return render_template("professors/groups_list.html", groups=groups, scores=scores)

@professors_bp.route("/groups/<int:group_id>/modal")
@login_required
@role_required("professor")
def group_detail_modal(group_id):
    g = _ensure_owns_group(group_id)
    students = (Student.query
                .join(GroupStudent, GroupStudent.student_id == Student.id)
                .filter(GroupStudent.group_id == g.id)
                .order_by(Student.name.asc())
                .all())
    by_inst = {r.instrument: r for r in GroupAssessment.query.filter_by(group_id=g.id).all()}
    ri  = by_inst.get(Instrument.RELATORIO_I)
    rii = by_inst.get(Instrument.RELATORIO_II)
    pp  = by_inst.get(Instrument.PAPER)
    return render_template("professors/_group_detail_modal.html",
                           group=g, students=students, ri=ri, rii=rii, paper=pp)

@professors_bp.route("/groups/<int:group_id>")
@login_required
@role_required("professor")
def group_detail(group_id):
    g = _ensure_owns_group(group_id)

    students = (Student.query
                .join(GroupStudent, GroupStudent.student_id == Student.id)
                .filter(GroupStudent.group_id == g.id)
                .order_by(Student.name.asc())
                .all())

    by_inst = _assessments_by_instrument(g.id)
    ri  = by_inst.get(Instrument.RELATORIO_I)
    rii = by_inst.get(Instrument.RELATORIO_II)
    pp  = by_inst.get(Instrument.PAPER)

    return render_template("professors/group_detail.html",
                           group=g, students=students,
                           ri=ri, rii=rii, paper=pp)

@professors_bp.route("/groups/<int:group_id>/grades", methods=["GET", "POST"])
@login_required
@role_required("professor")
def group_grades_edit(group_id):
    group = Group.query.get_or_404(group_id)
    # (se houver checagem de que o professor é o orientador, mantenha)

    form = GradeForm()
    if form.validate_on_submit():
        score_ri  = Decimal("0.5") if form.relatorio_i.data  else Decimal("0.0")
        score_rii = Decimal("0.5") if form.relatorio_ii.data else Decimal("0.0")
        paper_val = form.paper.data
        score_paper = Decimal(str(paper_val)) if paper_val is not None else Decimal("0.0")

        upsert_assessment(group.id, INST_RI,    score_ri,    current_user.id)
        upsert_assessment(group.id, INST_RII,   score_rii,   current_user.id)
        upsert_assessment(group.id, INST_PAPER, score_paper, current_user.id)

        db.session.commit()
        flash("Notas atualizadas.", "success")
        return redirect(url_for("professors.group_detail", group_id=group.id))

    # Pre-fill
    ri_score   = get_assessment_score(group.id, INST_RI)  or Decimal("0")
    rii_score  = get_assessment_score(group.id, INST_RII) or Decimal("0")
    paper_score= get_assessment_score(group.id, INST_PAPER) or Decimal("0")

    form.relatorio_i.data  = (ri_score  >= Decimal("0.5"))
    form.relatorio_ii.data = (rii_score >= Decimal("0.5"))
    form.paper.data = float(paper_score) if paper_score is not None else None

    return render_template("professors/grades_edit.html", group=group, form=form)
    
@professors_bp.route("/offerings/<int:offering_id>/export/csv", methods=["GET"])
@login_required
@role_required("professor")
def export_offering_csv(offering_id):
    off = Offering.query.get_or_404(offering_id)
    _assert_offering_access(off)

    rows = _collect_export_rows(off)
    headers = [
        "grupo", "rgm", "aluno", "campus", "oferta",
        "orientador", "relatorio_i", "relatorio_ii", "paper", "banner_media"
    ]

    si = StringIO(newline="")
    writer = csv.writer(si)
    writer.writerow(headers)
    for r in rows:
        out = []
        for val in r:
            out.append(f"{val:.2f}" if isinstance(val, float) else val)
        writer.writerow(out)

    fname = f"notas_oferta_{off.code}_{_stamp()}.csv"
    resp = Response(si.getvalue(), mimetype="text/csv; charset=utf-8")
    resp.headers["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp

@professors_bp.route("/offerings/<int:offering_id>/export/xlsx", methods=["GET"])
@login_required
@role_required("professor")
def export_offering_xlsx(offering_id):
    off = Offering.query.get_or_404(offering_id)
    _assert_offering_access(off)

    if openpyxl is None:
        flash("Para exportar Excel, instale o pacote 'openpyxl' (pip install openpyxl).", "warning")
        return redirect(url_for("professors.offerings_list"))

    rows = _collect_export_rows(off)
    headers = [
        "Nº do grupo", "RGM", "Aluno", "Campus", "Oferta",
        "Orientador", "Relatório I", "Relatório II", "Paper", "Apresentação de Banner (média)"
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notas"

    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    for r in rows:
        ws.append([round(v, 2) if isinstance(v, float) else v for v in r])

    from openpyxl.utils import get_column_letter
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    num_fmt = "0.00"
    for row in ws.iter_rows(min_row=2, min_col=7, max_col=10):
        for cell in row:
            if isinstance(cell.value, (float, int)):
                cell.number_format = num_fmt

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    fname = f"notas_oferta_{off.code}_{_stamp()}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@professors_bp.get("/offerings/<int:offering_id>")
@login_required
@role_required("professor")
def offering_detail(offering_id: int):
    off = Offering.query.get_or_404(offering_id)

    # Permissão: dono da oferta (ou admin)
    owner_id = getattr(off, "id_professor", None) or getattr(off, "professor_id", None)
    if owner_id != current_user.id and not getattr(current_user, "is_admin", False):
        abort(403)

    # -------- helpers p/ lidar com Instrument podendo ser Enum/str --------
    def _member(names):
        for n in names:
            if hasattr(Instrument, n):
                return getattr(Instrument, n)
        return None

    def _cond(possible_names):
        parts = []
        m = _member(possible_names)
        if m is not None:
            parts.append(GroupAssessment.instrument == m)
        parts.extend(GroupAssessment.instrument == s for s in possible_names)
        return or_(*parts)

    ri_cond    = _cond(["RI", "RELATORIO_I", "relatorio_i"])
    rii_cond   = _cond(["RII", "RELATORIO_II", "relatorio_ii"])
    paper_cond = _cond(["PAPER", "paper"])

    ri_col    = func.max(case((ri_cond, GroupAssessment.score), else_=None)).label("ri")
    rii_col   = func.max(case((rii_cond, GroupAssessment.score), else_=None)).label("rii")
    paper_col = func.max(case((paper_cond, GroupAssessment.score), else_=None)).label("paper")

    # -------------------- CONSULTA: partir de Student ---------------------
    q = (
        db.session.query(
            Student,          # aluno
            Group,            # grupo (pode ser None)
            User,             # orientador (pode ser None)
            ri_col, rii_col, paper_col,
        )
        .filter(Student.offering_id == off.id)
    )

    

    # Se for um Model GroupMember, use este bloco no lugar do bloco acima:
    q = (
        q.outerjoin(GroupStudent, GroupStudent.student_id == Student.id)
           .outerjoin(Group, GroupStudent.group_id == Group.id)
     )

    q = (
        q.outerjoin(User, Group.orientador_user_id == User.id)
         .outerjoin(GroupAssessment, GroupAssessment.group_id == Group.id)
         .group_by(Student.id, Group.id, User.id)
         .order_by(Student.name)
    )

    rows = q.all()

    def _f(x): return float(x) if x is not None else None
    def _banner(a, b, c):
        vals = [v for v in (a, b, c) if v is not None]
        return round(sum(vals)/len(vals), 2) if vals else None

    data = []
    for s, g, orient, ri, rii, paper in rows:
        ri, rii, paper = _f(ri), _f(rii), _f(paper)
        data.append(dict(
            student=s,
            group=g,
            orientador=orient,
            ri=ri, rii=rii, paper=paper,
            banner=_banner(ri, rii, paper),
        ))

    return render_template(
        "professors/offering_detail.html",
        off=off,
        rows=data,
        total=len(data),
    )


# Tenta carregar a tabela de avaliações de banner (se existir no seu projeto)
try:
    from ..models import BannerEvaluation  # ajuste se o nome do modelo for diferente
except Exception:
    BannerEvaluation = None

# Tenta carregar openpyxl para gerar .xlsx
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except Exception:
    openpyxl = None

def _assert_offering_access(off: Offering):
    """Garante que a oferta pertence ao professor logado (se o campo existir)."""
    prof_id = getattr(off, "professor_id", getattr(off, "id_professor", None))
    if prof_id is not None and prof_id != current_user.id:
        abort(403)

def _collect_export_rows(off: Offering):
    """
    Monta as linhas do relatório:
    [grupo_id, rgm, nome, campus, oferta, orientador, ri, rii, paper, banner]
    """
    # Alunos da oferta
    students = (
        Student.query.filter_by(offering_id=off.id)
        .options(joinedload(Student.campus), joinedload(Student.offering))
        .all()
    )
    if not students:
        return []

    student_ids = [s.id for s in students]

    # Mapeia aluno -> grupo
    gs_rows = GroupStudent.query.filter(GroupStudent.student_id.in_(student_ids)).all()
    student_group = {row.student_id: row.group_id for row in gs_rows}

    group_ids = sorted({row.group_id for row in gs_rows if row.group_id is not None})
    groups_by_id = {g.id: g for g in Group.query.filter(Group.id.in_(group_ids)).all()}

    # Carrega notas por grupo (RI, RII, PAPER)
    assessments = (
        GroupAssessment.query.filter(GroupAssessment.group_id.in_(group_ids)).all()
    )
    grades = {}
    for a in assessments:
        d = grades.setdefault(a.group_id, {})
        # Instrument enum -> rótulo curto
        name = str(a.instrument.name if hasattr(a.instrument, "name") else a.instrument)
        name = name.upper()
        if "I" in name and "II" not in name:
            key = "ri"
        elif "II" in name:
            key = "rii"
        elif "PAPER" in name:
            key = "paper"
        else:
            # desconhecido (ignora)
            continue
        d[key] = float(a.score) if a.score is not None else None

    # Média do banner por grupo (se existir modelo)
    banner_avg = {}
    if BannerEvaluation is not None:
        res = (
            db.session.query(
                BannerEvaluation.group_id, func.avg(BannerEvaluation.score)
            )
            .filter(BannerEvaluation.group_id.in_(group_ids))
            .group_by(BannerEvaluation.group_id)
            .all()
        )
        for gid, avg_ in res:
            banner_avg[gid] = float(avg_) if avg_ is not None else None

    # Monta linhas
    rows = []
    for s in students:
        gid = student_group.get(s.id)
        g = groups_by_id.get(gid) if gid else None

        # tenta extrair o orientador (nome) com fallback
        orientador_name = "-"
        if g is not None:
            orientador_obj = getattr(g, "orientador", None) or getattr(g, "advisor", None)
            if orientador_obj and isinstance(orientador_obj, User):
                orientador_name = orientador_obj.full_name
            else:
                # Se só tiver id (ex.: g.orientador_id), tente pegar 1x
                for attr_id in ("orientador_id", "advisor_id"):
                    if hasattr(g, attr_id):
                        o = User.query.get(getattr(g, attr_id))
                        if o:
                            orientador_name = o.full_name
                        break

        g_grades = grades.get(gid, {}) if gid else {}
        rows.append([
            gid or "-",                         # 1. nº do grupo
            s.rgm,                              # 2. rgm
            s.name,                             # 3. nome
            s.campus.name if s.campus else "-", # 4. campus
            s.offering.code if s.offering else "-",  # 5. oferta
            orientador_name,                    # 6. orientador
            g_grades.get("ri", None),           # 7. RI
            g_grades.get("rii", None),          # 8. RII
            g_grades.get("paper", None),        # 9. Paper
            (None if gid is None else banner_avg.get(gid)),  # 10. Banner (média)
        ])
    return rows