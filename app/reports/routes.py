from datetime import datetime
from decimal import Decimal
from io import StringIO, BytesIO
import csv
from collections import defaultdict

from flask import request, Response, send_file, redirect, url_for, flash
from flask_login import login_required, current_user

from app.reports import reports_bp
from app.utils.decorators import role_required
from app.extensions import db
from app.models import (
    Student, GroupStudent, Group, GroupAssessment, BannerEvaluation,
    Offering, User
)
from sqlalchemy.orm import joinedload
from sqlalchemy import func

# openpyxl é opcional
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except Exception:
    openpyxl = None


def _stamp() -> str:
    return datetime.now().strftime("%Y%m%d-%H%M%S")


def _offerings_for_current_prof():
    """
    Retorna as ofertas do professor logado, considerando seu schema
    (já vi 'professor_id' e 'id_professor' no seu projeto).
    """
    q = Offering.query
    if hasattr(Offering, "professor_id"):
        q = q.filter(Offering.professor_id == current_user.id)
    elif hasattr(Offering, "id_professor"):
        q = q.filter(Offering.id_professor == current_user.id)
    else:
        # Caso o modelo não tenha esses campos, retorna vazio
        return []
    return q.all()


def _collect_rows(off_ids):
    """
    Monta linhas no mesmo formato do export do professor:
    [grupo, rgm, aluno, campus, oferta, orientador, ri, rii, paper, banner_media]
    """
    if not off_ids:
        return []

    # 1) alunos das ofertas
    students = (
        Student.query
        .filter(Student.offering_id.in_(off_ids))
        .options(joinedload(Student.campus), joinedload(Student.offering))
        .all()
    )
    if not students:
        return []

    student_ids = [s.id for s in students]

    # 2) mapeia aluno -> grupo
    gs_rows = GroupStudent.query.filter(GroupStudent.student_id.in_(student_ids)).all()
    student_group = {row.student_id: row.group_id for row in gs_rows}

    group_ids = sorted({row.group_id for row in gs_rows if row.group_id is not None})
    groups_by_id = {g.id: g for g in Group.query.filter(Group.id.in_(group_ids)).all()}

        # 3) notas por grupo (RI / RII / PAPER), tolerando Enum/strings
    assessments = GroupAssessment.query.filter(GroupAssessment.group_id.in_(group_ids)).all()
    grades = {}
    for a in assessments:
        d = grades.setdefault(a.group_id, {})
        key = _instrument_key(a.instrument)
        if not key:
            continue
        d[key] = float(a.score) if a.score is not None else None


    # 4) média do banner (se existir)
    banner_avg = {}
    if group_ids and BannerEvaluation is not None:
        res = (
            db.session.query(BannerEvaluation.group_id, func.avg(BannerEvaluation.score))
            .filter(BannerEvaluation.group_id.in_(group_ids))
            .group_by(BannerEvaluation.group_id)
            .all()
        )
        for gid, avg_ in res:
            banner_avg[gid] = float(avg_) if avg_ is not None else None

    # 5) linhas
    rows = []
    for s in students:
        gid = student_group.get(s.id)
        g = groups_by_id.get(gid) if gid else None

        # orientador (tolerante)
        orientador_name = "-"
        if g is not None:
            orient_obj = getattr(g, "orientador", None) or getattr(g, "advisor", None)
            if isinstance(orient_obj, User):
                orientador_name = orient_obj.full_name
            elif hasattr(g, "orientador_user_id") and g.orientador_user_id:
                u = User.query.get(g.orientador_user_id)
                if u:
                    orientador_name = u.full_name

        g_grades = grades.get(gid, {}) if gid else {}
        rows.append([
            gid or "-",                                  # grupo
            s.rgm,                                       # rgm
            s.name,                                      # aluno
            s.campus.name if s.campus else "-",          # campus
            s.offering.code if s.offering else "-",      # oferta
            orientador_name,                             # orientador
            g_grades.get("ri"),                          # RI
            g_grades.get("rii"),                         # RII
            g_grades.get("paper"),                       # Paper
            (None if gid is None else banner_avg.get(gid)),  # média banner
        ])
    return rows


@reports_bp.get("/export")
@login_required
@role_required("admin", "professor")
def export():
    """
    Admin: exporta TODOS os alunos.
    Professor: exporta SOMENTE as suas ofertas.
    """
    fmt = (request.args.get("fmt") or "csv").lower()

    if getattr(current_user, "role_value", None) == "professor":
        offs = _offerings_for_current_prof()
        off_ids = [o.id for o in offs]
    else:
        # admin -> todas
        off_ids = [o.id for o in Offering.query.all()]

    rows = _collect_rows(off_ids)

    # CSV
    if fmt == "csv":
        headers = [
            "grupo", "rgm", "aluno", "campus", "oferta",
            "orientador", "relatorio_i", "relatorio_ii", "paper", "banner_media"
        ]
        si = StringIO(newline="")
        w = csv.writer(si)
        w.writerow(headers)
        for r in rows:
            w.writerow([f"{v:.2f}" if isinstance(v, float) else v for v in r])
        fname = f"relatorio_{_stamp()}.csv"
        resp = Response(si.getvalue(), mimetype="text/csv; charset=utf-8")
        resp.headers["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp

    # XLSX
    if fmt == "xlsx":
        if openpyxl is None:
            flash("Para exportar Excel, instale 'openpyxl' (pip install openpyxl).", "warning")
            # volta para um lugar seguro
            return redirect(url_for("admin.dashboard") if getattr(current_user, "role_value", "") == "admin"
                            else url_for("professors.offerings_list"))

        headers = [
            "Nº do grupo", "RGM", "Aluno", "Campus", "Oferta",
            "Orientador", "Relatório I", "Relatório II", "Paper", "Apresentação de Banner (média)"
        ]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório"
        ws.append(headers)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        for r in rows:
            ws.append([round(v, 2) if isinstance(v, float) else v for v in r])

        from openpyxl.utils import get_column_letter
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 22

        for row in ws.iter_rows(min_row=2, min_col=7, max_col=10):
            for cell in row:
                if isinstance(cell.value, (float, int)):
                    cell.number_format = "0.00"

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        fname = f"relatorio_{_stamp()}.xlsx"
        return send_file(
            bio,
            as_attachment=True,
            download_name=fname,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # formarto desconhecido
    flash("Formato inválido. Use ?fmt=csv ou ?fmt=xlsx.", "warning")
    return redirect(url_for("admin.dashboard") if getattr(current_user, "role_value", "") == "admin"
                    else url_for("professors.offerings_list"))

# app/reports/routes.py

def _instrument_key(inst) -> str | None:
    """
    Normaliza o nome do instrumento para uma das chaves: 'ri' | 'rii' | 'paper'.
    Aceita Enum (usa .name) ou string; tolera underscore/espaco e acentos.
    """
    name = getattr(inst, "name", inst)
    if not name:
        return None
    name = str(name).upper()

    # normaliza acentos e separadores
    name = (name
            .replace("Á", "A").replace("À", "A").replace("Â", "A").replace("Ã", "A")
            .replace("É", "E").replace("Ê", "E")
            .replace("Í", "I")
            .replace("Ó", "O").replace("Ô", "O")
            .replace("Ú", "U")
            .replace("-", " ").replace("_", " ").strip())

    # matches mais comuns
    if name in {"RI", "RELATORIO I", "RELATORIO I"}:
        return "ri"
    if name in {"RII", "RELATORIO II"}:
        return "rii"
    if "PAPER" in name:
        return "paper"
    return None

@reports_bp.get("/groups.<fmt>")
@login_required
@role_required("admin", "professor")
def groups_export(fmt: str):
    """
    Exporta grupos em colunas achatadas:
      # Grupo, Título, Orientador, nome_1, rgm_1, nome_2, rgm_2, ...
    - admin: todos os grupos
    - professor: apenas os grupos que orienta
    Parâmetros:
      - fmt: 'csv' | 'xlsx'
      - ?max=N  -> força o número de pares nome/rgm
    """
    # 1) Carrega grupos (filtra por professor se necessário)
    base = Group.query.options(joinedload(Group.orientador)).order_by(Group.id.asc())

    role_val = getattr(current_user, "role_value", None) or (
        current_user.role.value if hasattr(current_user.role, "value") else current_user.role
    )
    role_val = (role_val or "").lower()
    if role_val == "professor":
        base = base.filter(Group.orientador_user_id == current_user.id)

    groups = base.all()
    group_ids = [g.id for g in groups]

    # 2) Pré-carrega membros
    members_by_group: dict[int, list[Student]] = defaultdict(list)
    if group_ids:
        links = (
            GroupStudent.query
            .filter(GroupStudent.group_id.in_(group_ids))
            .options(joinedload(GroupStudent.student))
            .all()
        )
        for gs in links:
            if gs.student:
                members_by_group[gs.group_id].append(gs.student)

    # 3) Decide quantos pares nome/rgm exportar
    max_param = request.args.get("max", type=int)
    if max_param and max_param > 0:
        max_members = max_param
    else:
        # calcula automaticamente com piso mínimo de 3
        max_members = max([len(members_by_group.get(g.id, [])) for g in groups] + [3])

    # 4) Monta header achatado
    header = ["# Grupo", "Título do trabalho", "Orientador"]
    for i in range(1, max_members + 1):
        header += [f"nome_{i}", f"rgm_{i}"]

    # 5) Linhas
    out_rows = []
    for g in groups:
        orient = g.orientador.full_name if getattr(g, "orientador", None) else "-"
        students = sorted(members_by_group.get(g.id, []), key=lambda s: (s.name or "").upper())
        # corta/ completa
        pairs = [(s.name or "", s.rgm or "") for s in students[:max_members]]
        while len(pairs) < max_members:
            pairs.append(("", ""))
        flat = []
        for n, r in pairs:
            flat += [n, r]
        out_rows.append([g.id, (g.title or "-"), orient, *flat])

    # 6) XLSX
    if fmt.lower() == "xlsx":
        if openpyxl is None:
            flash("Para exportar Excel, instale 'openpyxl' (pip install openpyxl).", "warning")
            return redirect(url_for("admin.groups_list"))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Grupos"
        ws.append(header)
        for c in ws[1]:
            c.font = Font(bold=True)

        for row in out_rows:
            ws.append(row)

        # larguras agradáveis
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, len(header) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 22

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        fname = f"grupos_{_stamp()}.xlsx"
        return send_file(
            bio,
            as_attachment=True,
            download_name=fname,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # 7) CSV (UTF-8 BOM; separador padrão “,” — mude para ';' se preferir)
    si = StringIO(newline="")
    writer = csv.writer(si)  # csv.writer(si, delimiter=';') se quiser ponto-e-vírgula
    writer.writerow(header)
    writer.writerows(out_rows)
    fname = f"grupos_{_stamp()}.csv"
    resp = Response(si.getvalue().encode("utf-8-sig"), mimetype="text/csv; charset=utf-8")
    resp.headers["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp

